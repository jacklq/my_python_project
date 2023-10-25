import sys

from docxtpl import DocxTemplate
from openpyxl import load_workbook
import os
import win32api
import win32print
import yaml

# 定义一个空的列表用于存放数据
contexts = []
new_word_path = ""
max_num_new_word = 0  # 每次做多生成打印0个新word
template_path = ""
registration_form_path = ""


#调用打印机
def printer_loading(filename):
    print("开始调用打印机")
    win32api.ShellExecute(
            0,  # 父窗口的句柄，如果没有父窗口，则为0
            "print",  # 要进行的操作，为“open”、“print”或者为空
            filename,  # 要打印的文件
            win32print.GetDefaultPrinter(),  # 要向程序传递的参数.这里调用系统默认的打印机
            ".",  # 程序初始化的目录
            0  # 是否显示窗口
    )
    print("调用打印机成功")
    # 获取真实excle行数
def get_max_row(sheet):
    i=sheet.max_row
    real_max_row = 0
    while i > 0:
        row_dict = {i.value for i in sheet[i]}
        if row_dict == {None}:
            i = i-1
        else:
            real_max_row = i
            break

    return real_max_row


# 获取excel中的数据
def get_excel_data(excel_pathname):
    print("开始获取excel中的数据")
    # 获取excel句柄
    wk = load_workbook(excel_pathname)
    # 获取excel对应的表单
    ws = wk['收文登记']
    real_max_row=get_max_row(ws)
    # Start row, where data begins
    # Getting cell value, from columns A, B and C
    # Iterating through rows 2, 3, 4 ...
    if real_max_row > max_num_new_word:
        for row in range(real_max_row - max_num_new_word + 1, real_max_row + 1):
            name = ws["B" + str(row)].value
            file_num = ws["C" + str(row)].value
            source = ws["D" + str(row)].value
            time = str(ws["E" + str(row)].value)
            time_year = time[0:4]
            time_mon = time[4:6]
            time_day = time[6:8]
            secret = ws["F" + str(row)].value
            suggestion = ws["G" + str(row)].value
            context = {"name": name, "file_num": file_num,"source": source,
                       "time": time, "time_year": time_year,"time_mon": int(time_mon), "time_day": int(time_day),
                       "secret": secret,"suggestion": suggestion}
            contexts.append(context)  # 将每条字典，存到contexts数组中
    else:
        row = real_max_row
        name = ws["B" + str(row)].value
        file_num = ws["C" + str(row)].value
        source = ws["D" + str(row)].value
        time = str(ws["E" + str(row)].value)
        time_year = time[0:4]
        time_mon = time[4:6]
        time_day = time[6:8]
        secret = ws["F" + str(row)].value
        suggestion = ws["G" + str(row)].value
        context = {"name": name, "file_num": file_num, "source": source,
                   "time": time, "time_year": time_year, "time_mon": time_mon, "time_day": time_day,
                   "secret": secret, "suggestion": suggestion}
        contexts.append(context)
    print("获取excel中的数据成功")
    return contexts


def finish_word_data(word_pathname, contexts):
    if not os.path.exists(new_word_path):
        # 创建要保存的文件夹
        os.mkdir(new_word_path)
    print("开始生成word文档")
    for context in contexts:
        print("开始生成word文档")
        print(context)
        newWordName = (new_word_path + "/{}.docx").format(context["time"] + " " + context["name"])
        if not os.path.exists(newWordName):
            tpl = DocxTemplate(word_pathname)  # 获取原有的word模板数据
            tpl.render(context)  # 利用插值表达式{{ example }}  替换word模板中需要替换的地方
            tpl.save(newWordName)  # 保存为新的word，修改后的
            print("生成word文档成功")
            # 打印新word
            printer_loading(newWordName)
        else:
            print("word文档已存在，已打印过，不再进行重复打印")

if __name__ == "__main__":
    sys.stderr = open('error_log.txt', 'w')
    sys.stdout = open('run_log.txt', 'w')
    with open('config.yaml', 'r', encoding="utf-8") as f:  # 用with读取文件更好
        configs = yaml.load(f, Loader=yaml.FullLoader)  # 按字典格式读取并返回
        # 显示读取后的内容
    new_word_path = configs["new_word_path"]
    max_num_new_word = configs["max_num_new_word"]  # 每次做多生成打印n个新word
    template_path = configs["template_path"]
    registration_form_path = configs["registration_form_path"]


    contexts = get_excel_data(registration_form_path)
    finish_word_data(template_path, contexts)

    # path = 'D:\softwareProfessionInstall\pycharm\pythonProject\\2023阅办单'  # 你要批量打印文件的路径
    # # 指定另一个打印机名作为默认打印机
    # # win32print.SetDefaultPrinter('HP Color MFP E87640-50-60 PCL-6 (V4) (网络)')  # 这里可以换成其他打印机名称
    # print("Your system default printer name is:", win32print.GetDefaultPrinter())  # 识别到你的系统默认打印机
    # print(os.listdir(path))
    # for file in os.listdir(path):
    #     printer_loading(os.path.join(path, file))
